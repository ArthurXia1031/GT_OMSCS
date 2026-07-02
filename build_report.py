#!/usr/bin/env python3
"""
build_report.py — Inject a CMPT rule-performance Excel workbook into the dashboard HTML.

Model A data channel (no manual upload, no browser-side .xlsx parsing):

    your existing Excel  ->  openpyxl reads it  ->  JSON  ->  baked into HTML template

The HTML template (cmpt_report_pro.html) already contains the hook:

    // __INJECTED_DATA_PLACEHOLDER__
    const INJECTED_DATA = null;

This script replaces `const INJECTED_DATA = null;` with the real object. If the template
is left untouched (INJECTED_DATA === null), the report falls back to its synthetic demo data.

IMPORTANT: the script keys off COLUMN HEADERS and SHEET NAMES, not the actual values, and
matches them tolerantly (case-insensitive, punctuation-insensitive, contains-fallback).
So once it works against the sample workbook here, it works against your real internal
workbook as long as the headers/sheets look like the ones in the SAS export.

Usage
-----
    # validate first — prints what it found vs what it expects, changes nothing:
    python build_report.py --excel CMPT_6m_report.xlsx --template cmpt_report_pro.html --validate

    # generate the report:
    python build_report.py --excel CMPT_6m_report.xlsx \
                           --template cmpt_report_pro.html \
                           --out cmpt_report_$(date +%Y%m%d).html

Dependencies: openpyxl  (pip install openpyxl)
"""

import argparse
import datetime as dt
import json
import os
import re
import sys

try:
    from openpyxl import load_workbook
except ImportError:
    sys.exit("ERROR: openpyxl is required.  Run:  pip install openpyxl")

PLACEHOLDER = "const INJECTED_DATA = null;"

# ---------------------------------------------------------------------------
# Configuration — the only thing you may need to touch if your headers differ.
# Each canonical field maps to a list of accepted header spellings (matched
# case/punctuation-insensitively, with a "contains" fallback).
# ---------------------------------------------------------------------------
COLUMN_ALIASES = {
    "id":      ["RULE_ID", "RULEID", "ID"],
    "name":    ["RULE_NAME", "RULENAME", "NAME"],
    "status":  ["RULE_STATUS", "STATUS"],
    "decision":["RULE_DECISION", "DECISION"],
    "loc":     ["RULE_LOCATION", "LOCATION", "LOC"],
    "month":   ["AUTH_MONTH", "AUTHMONTH", "MONTH"],
    "totAcct": ["FINAL_TOT_ACCT", "TOT_ACCT", "TOTACCT", "TOTAL_ACCT"],
    "frdAcct": ["FINAL_FRD_ACCT", "FRD_ACCT", "FRDACCT", "FRAUD_ACCT"],
    "totAuth": ["FINAL_TOT_AUTH", "TOT_AUTH", "TOTAUTH", "TOTAL_AUTH"],
    "frdAuth": ["FINAL_FRD_AUTH", "FRD_AUTH", "FRDAUTH", "FRAUD_AUTH"],
}
STRING_FIELDS = {"name", "status", "loc", "month"}
NUMERIC_FIELDS = {"totAcct", "frdAcct", "totAuth", "frdAuth"}

# The 6 dashboard sub-tabs (view keys used by the HTML) and how to recognise the
# matching sheet from its name: (must-contain-all tokens after normalisation).
VIEW_SHEET_RULES = {
    "cmpt_ct_30d":   {"scope": "CMPT", "metric": "CT",      "window": "30"},
    "cmpt_ct_6m":    {"scope": "CMPT", "metric": "CT",      "window": "6M"},
    "all_ct_6m":     {"scope": "ALL",  "metric": "CT",      "window": "6M"},
    "cmpt_trig_30d": {"scope": "CMPT", "metric": "TRIGGER", "window": "30"},
    "cmpt_trig_6m":  {"scope": "CMPT", "metric": "TRIGGER", "window": "6M"},
    "all_trig_6m":   {"scope": "ALL",  "metric": "TRIGGER", "window": "6M"},
}


def norm(s):
    """Uppercase, drop everything but A-Z 0-9 — for tolerant matching."""
    return re.sub(r"[^A-Z0-9]", "", str(s).upper())


def coerce_num(v):
    if v is None:
        return 0
    if isinstance(v, (int, float)):
        return v
    if isinstance(v, (dt.datetime, dt.date)):
        return 0
    s = re.sub(r"[^0-9.\-]", "", str(v))
    try:
        f = float(s)
        return int(f) if f.is_integer() else f
    except ValueError:
        return 0


def cell_str(v):
    """JSON-safe string for any cell. Excel dates come back as datetime → format compactly
    (e.g. '20Mar2026', matching the AUTH_MONTH style); everything else via str()."""
    if v in (None, ""):
        return ""
    if isinstance(v, (dt.datetime, dt.date)):
        return v.strftime("%d%b%Y")
    return str(v).strip()


def json_safe(v):
    """Make any raw cell value JSON-serializable (datetimes → string, keep numbers as-is)."""
    if isinstance(v, (dt.datetime, dt.date)):
        return v.strftime("%d%b%Y")
    if isinstance(v, (int, float, str)) or v is None:
        return "" if v is None else v
    return str(v)


def read_sheet(ws):
    """Return (headers, list-of-dict-rows). First non-empty row = header row."""
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []
    # find header row (first row with >=2 non-empty cells)
    h_idx = 0
    for i, r in enumerate(rows):
        if sum(1 for c in r if c not in (None, "")) >= 2:
            h_idx = i
            break
    headers = [("" if c is None else str(c).strip()) for c in rows[h_idx]]
    data = []
    for r in rows[h_idx + 1:]:
        if all(c in (None, "") for c in r):
            continue
        data.append({headers[i]: r[i] for i in range(len(headers)) if headers[i]})
    return headers, data


def resolve_columns(headers):
    """Map each canonical field -> actual header string (or None if absent)."""
    nmap = {norm(h): h for h in headers if h}
    resolved = {}
    for field, aliases in COLUMN_ALIASES.items():
        hit = None
        for a in aliases:                       # exact normalised match first
            if norm(a) in nmap:
                hit = nmap[norm(a)]
                break
        if not hit:                              # contains-fallback
            for na in (norm(a) for a in aliases):
                for nh, h in nmap.items():
                    if na and na in nh:
                        hit = h
                        break
                if hit:
                    break
        resolved[field] = hit
    return resolved


def sheet_to_view(sheet_name):
    """Return the canonical view key for a sheet name, or None if it's not a rule sheet."""
    n = norm(sheet_name)
    # Direct match: sheet literally named like the view key (e.g. mock "cmpt_trig_6m").
    for view in VIEW_SHEET_RULES:
        if norm(view) == n:
            return view
    if "RULE" not in n:
        return None
    is_all = n.startswith("ALL")
    is_trig = "TRIG" in n
    is_6m = "6M" in n
    is_30 = "30" in n
    if not (is_6m or is_30):
        return None
    scope = "ALL" if is_all else "CMPT"
    metric = "TRIGGER" if is_trig else "CT"
    window = "6M" if is_6m else "30"
    for view, rule in VIEW_SHEET_RULES.items():
        if rule["scope"] == scope and rule["metric"] == metric and rule["window"] == window:
            return view
    return None


def map_rule_rows(data_rows, cols, max_rows):
    out = []
    for i, row in enumerate(data_rows[:max_rows]):
        rec = {}
        for field, header in cols.items():
            raw = row.get(header) if header else None
            if field in NUMERIC_FIELDS:
                rec[field] = coerce_num(raw)
            elif field == "status":
                rec[field] = (str(raw).strip().upper()[:1] if raw not in (None, "") else "E")
            elif field == "id":
                rec[field] = (cell_str(raw) if raw not in (None, "") else 90000 + i)
            else:
                rec[field] = cell_str(raw)
        out.append(rec)
    return out


# ===========================================================================
# §04 VOLUME MONITOR — CPP batch volume from Snowflake (no CSV; rows → JSON).
# Query output contract: 3 columns  SOURCE (str) | BATCH_DT (date) | TOT_ACCT (number),
# one row per source per day, ~6 months ending today.
# ===========================================================================
NDAYS_VOL = 45   # dashboard §04 renders a 45-day window; we keep the most recent 45 days.

# ── Snowflake connection settings (from your notebook) — edit if your env differs ──

USE_DFS_PROXY    = True           # set False off-network

# >>> PASTE each §04 group's batch-volume SQL here (3 cols: SOURCE str | BATCH_DT date | TOT_ACCT number) <<<
CPP_VOLUME_SQL = """
SELECT SOURCE, BATCH_DT, TOT_ACCT
FROM   <YOUR_CPP_VOLUME_TABLE>            -- <<< replace with your real table / view / query
WHERE  BATCH_DT >= DATEADD('month', -6, CURRENT_DATE())
ORDER  BY SOURCE, BATCH_DT
"""

TESTING_VOLUME_SQL = """
SELECT SOURCE, BATCH_DT, TOT_ACCT
FROM   <YOUR_TESTING_VOLUME_TABLE>        -- <<< replace with your real Testing query
WHERE  BATCH_DT >= DATEADD('month', -6, CURRENT_DATE())
ORDER  BY SOURCE, BATCH_DT
"""

# Merchant LUL — TOT_ACCT here is a MERCHANT count (unit shown as "merchants" in §04).
MLIST_VOLUME_SQL = """
SELECT SOURCE, BATCH_DT, TOT_ACCT
FROM   <YOUR_MERCHANT_LUL_VOLUME_TABLE>   -- <<< replace with your real Merchant-LUL query
WHERE  BATCH_DT >= DATEADD('month', -6, CURRENT_DATE())
ORDER  BY SOURCE, BATCH_DT
"""

# §04 group -> SQL.  Fill in the ones you have; any query still containing "<YOUR_" is auto-skipped.
# Add "vendor" here the same way later.
VOLUME_SQL = {"cpp": CPP_VOLUME_SQL, "testing": TESTING_VOLUME_SQL, "mlist": MLIST_VOLUME_SQL}

# ── Q2: process daily metrics → feeds §00 Command Center, §01 Quality, §01.2 AFPR, §02 Trend, §05 ──
# ONE SQL PER FAMILY (like volume). Each query's result set columns (tolerant aliases):
#   <SRCE>  : CPP_SRCE / TESTING_SRCE / SOURCE / *_SRCE   (process name)
#   TIER_INFO (or TIER) : 'All' + 'T1'/'T2'/'T3'  ('All' row is used for the process-level series;
#                          T1/T2/T3 rows drive the tier breakdown)
#   AUTH_DT (or BATCH_DT) : date, ~last 90 days, daily
#   TOT_ACCT | FRAUD_ACCT (or FRD_ACCT) | FRAUD_DLR (or APRV_FRAUD_DLR)
# Extra columns (TOT_TRAN, FRAUD_TRAN, TOT_DLR, TFPR, AFPR, DFPR) are ignored — AFPR is recomputed.
CPP_PROCESS_SQL = """
<PASTE your CPP process-metrics SQL here — returns CPP_SRCE, TIER_INFO, AUTH_DT, TOT_ACCT, FRAUD_ACCT, FRAUD_DLR (and more)>
"""
TESTING_PROCESS_SQL = """
<YOUR_TESTING_PROCESS_SQL>
"""
VENDOR_PROCESS_SQL = """
<YOUR_VENDOR_PROCESS_SQL>
"""
# family -> SQL.  Fill the ones you have; any query still containing "<YOUR_" or "<PASTE" is skipped.
PROCESS_SQL = {"cpp": CPP_PROCESS_SQL, "testing": TESTING_PROCESS_SQL, "vendor": VENDOR_PROCESS_SQL}


def snowflake_connect(creds_path=None):
    """Open a Snowflake connection using a local cred JSON (username/password/role_id),
    mirroring the notebook pattern. Returns a live connection (warehouse/db/schema set)."""
    import snowflake.connector  # lazy import — only needed with --snowflake
    if USE_DFS_PROXY:
        os.environ["http_proxy"]  = ""
        os.environ["https_proxy"] = ""
        os.environ["no_proxy"]    = ""

    path = creds_path or CRED_PATH
    with open(path, "r") as openfile:
        creds = json.load(openfile)

    con = snowflake.connector.connect(
        user=creds["username"],
        password=creds["password"],
        authenticator=SF_AUTHENTICATOR,
        account=SF_ACCOUNT,
        role=creds["role_id"],
    )
    cur = con.cursor()
    cur.execute(f"USE warehouse {SF_WAREHOUSE}")
    cur.execute(f"USE database {SF_DATABASE}")
    cur.execute(f"USE schema {SF_SCHEMA}")
    cur.close()
    return con


def fetch_cpp_volume(conn, sql=CPP_VOLUME_SQL):
    """Run the CPP volume query → list of {SOURCE, BATCH_DT, TOT_ACCT} dicts.
    Uses the cursor directly (no pandas dependency); swap in pd.read_sql if you prefer."""
    # cursor.execute() runs ONE statement. Strip trailing ';' / blank lines so a copy-pasted
    # query that ends in a semicolon doesn't trip "statement count 2 != 1".
    cleaned = sql.strip()
    while cleaned.endswith(";"):
        cleaned = cleaned[:-1].strip()
    cur = conn.cursor()
    try:
        cur.execute(cleaned)
        cols = [c[0].upper() for c in cur.description]
        return [dict(zip(cols, row)) for row in cur.fetchall()]
    finally:
        cur.close()


def parse_date(v):
    if v is None:
        return None
    if isinstance(v, dt.datetime):
        return v.date()
    if isinstance(v, dt.date):
        return v
    s = str(v).strip()[:10]
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%Y/%m/%d", "%d%b%Y"):
        try:
            return dt.datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def build_volume_payload(rows, group="cpp", ndays=NDAYS_VOL):
    """Pure function: long rows (SOURCE,BATCH_DT,TOT_ACCT) → {group:{dates:[ndays], procs:[{id,short,series:[ndays]}]}}.
    Aligns every source to one contiguous ndays axis ending at the latest date; missing days → 0."""
    recs = []
    for r in rows:
        rl = {str(k).upper(): v for k, v in r.items()}
        src = str(rl.get("SOURCE", "")).strip()
        d = parse_date(rl.get("BATCH_DT"))
        v = coerce_num(rl.get("TOT_ACCT"))
        if src and d is not None:
            recs.append((src, d, v))
    if not recs:
        return None
    max_d = max(d for _, d, _ in recs)
    axis = [max_d - dt.timedelta(days=ndays - 1 - i) for i in range(ndays)]  # oldest → newest
    by_src = {}
    for src, d, v in recs:
        by_src.setdefault(src, {})[d] = v
    procs, aggregate = [], None
    for src in sorted(by_src):
        dmap = by_src[src]
        series = [int(round(dmap.get(day, 0))) for day in axis]
        # A SOURCE literally named ALL/TOTAL = the SQL-computed cross-source DISTINCT total;
        # use it as the group's aggregate row instead of summing per-source (which double-counts
        # accounts that are in more than one process). Don't list it as a process.
        if src.strip().upper() in ("ALL", "TOTAL", "ALL_SOURCES"):
            aggregate = series
        else:
            procs.append({"id": src, "short": src, "series": series})
    procs.sort(key=lambda p: p["series"][-1], reverse=True)   # head-heavy
    dates = ["%d/%d" % (d.month, d.day) for d in axis]
    out = {"dates": dates, "procs": procs}
    if aggregate is not None:
        out["aggregate"] = aggregate                          # deduped cross-source daily total
    return {group: out}


def _norm_family(v):
    u = str(v).upper()
    if "CPP" in u:
        return "cpp"
    if "VEND" in u:
        return "vendor"
    if "TEST" in u or "UDV" in u or "ATM" in u:
        return "testing"
    return None


def _norm_tier(v):
    s = str(v).strip().upper()
    if s in ("ALL", "ALLTIER", "ALL TIER", "ALL_TIER", "TOTAL"):
        return "ALL"
    if s.startswith(("1", "A")):
        return "T1"
    if s.startswith(("2", "B")):
        return "T2"
    if s.startswith(("3", "C")):
        return "T3"
    return s if s in ("T1", "T2", "T3") else "T3"


def _pick(rl, *names):
    """Value for the first matching column: exact name, else a key ending in the name (e.g. *_SRCE)."""
    for n in names:
        if n in rl:
            return rl[n]
    for n in names:
        for k in rl:
            if k.endswith(n):
                return rl[k]
    return None


def build_family_payload(rows, family, ndays=NDAYS_VOL, roll=15):
    """One family's Q2 rows → {family:{dates:[ndays], procs:[{id,short,afprReal,totReal,frdReal,dlrReal,tierDlr,tiers}]}}.
    Fraud reporting lags, so every displayed day is a TRAILING `roll`-day aggregate (T-roll+1 … T):
    that day's TOT/FRD/$ = sum over the prior `roll` days, AFPR = ΣTOT/ΣFRD over the window.
    Process-level series come from the 'All' tier row; T1/T2/T3 rows drive the tier breakdown.
    Column names tolerant: CPP_SRCE/*_SRCE/SOURCE, TIER_INFO/TIER, AUTH_DT/BATCH_DT,
    TOT_ACCT, FRAUD_ACCT/FRD_ACCT, FRAUD_DLR/APRV_FRAUD_DLR."""
    # Optional columns (drive §02 filters). If a column is entirely absent we emit NO field for it,
    # so the dashboard shows "data not available" instead of a made-up constant.
    #   FRAUD_DLR_HS   (All row)   -> High-Spender $ share            -> procs[].dlrRealHs
    #   FRAUD_DLR_V15  (tier rows) -> 15-min-verified $ (T2/T3)       -> procs[].tierDlrV15
    #   FRAUD_ACCT     (tier rows) -> tier fraud accounts (already present) -> procs[].tierFrd
    recs = []
    has_hs = False
    has_v15 = False
    for r in rows:
        rl = {str(k).upper(): v for k, v in r.items()}
        src = str(_pick(rl, "SOURCE", "SRCE") or "").strip()
        d = parse_date(_pick(rl, "AUTH_DT", "BATCH_DT", "DT"))
        if not (src and d is not None):
            continue
        hs_raw = _pick(rl, "FRAUD_DLR_HS", "FRD_DLR_HS", "APRV_FRAUD_DLR_HS")
        v15_raw = _pick(rl, "FRAUD_DLR_V15", "FRD_DLR_V15", "APRV_FRAUD_DLR_V15")
        if hs_raw is not None:
            has_hs = True
        if v15_raw is not None:
            has_v15 = True
        recs.append((src, d, _norm_tier(_pick(rl, "TIER_INFO", "TIER")),
                     coerce_num(_pick(rl, "TOT_ACCT")),
                     coerce_num(_pick(rl, "FRAUD_ACCT", "FRD_ACCT")),
                     coerce_num(_pick(rl, "FRAUD_DLR", "APRV_FRAUD_DLR")),
                     coerce_num(hs_raw), coerce_num(v15_raw)))
    if not recs:
        return None
    max_d = max(x[1] for x in recs)
    axis = [max_d - dt.timedelta(days=ndays - 1 - i) for i in range(ndays)]   # display days (oldest→newest)
    dates = ["%d/%d" % (d.month, d.day) for d in axis]

    # Raw daily data keyed by date (need a wider range than the axis to roll the earliest days).
    # all/tsum rows carry [tot,frd,dlr,hs]; ttot carries [tot,frd,dlr]; tv15 is a per-tier {date:$}.
    data = {}
    for src, d, tier, tot, frd, dlr, dhs, dv15 in recs:
        e = data.setdefault(src, {"has_all": False, "all": {}, "tsum": {},
                                  "tier": {"T1": {}, "T2": {}, "T3": {}},
                                  "ttot": {"T1": {}, "T2": {}, "T3": {}},
                                  "tv15": {"T1": {}, "T2": {}, "T3": {}}})
        if tier == "ALL":
            e["has_all"] = True
            a = e["all"].setdefault(d, [0, 0, 0, 0]); a[0] += tot; a[1] += frd; a[2] += dlr; a[3] += dhs
        elif tier in ("T1", "T2", "T3"):
            s = e["tsum"].setdefault(d, [0, 0, 0, 0]); s[0] += tot; s[1] += frd; s[2] += dlr; s[3] += dhs
            e["tier"][tier][d] = e["tier"][tier].get(d, 0) + dlr
            tt = e["ttot"][tier].setdefault(d, [0, 0, 0]); tt[0] += tot; tt[1] += frd; tt[2] += dlr
            e["tv15"][tier][d] = e["tv15"][tier].get(d, 0) + dv15

    def wsum(daymap, end_day, sel):
        """Trailing `roll`-day sum ending at end_day; sel picks the field (or identity for scalars)."""
        tot = 0
        for k in range(roll):
            v = daymap.get(end_day - dt.timedelta(days=k))
            if v is not None:
                tot += (v[sel] if isinstance(v, list) else v)
        return tot

    procs = []
    for src, e in data.items():
        pmap = e["all"] if e["has_all"] else e["tsum"]      # process-level raw daily
        tot_r, frd_r, dlr_r, afpr_r, hs_r = [], [], [], [], []
        for d in axis:
            t = wsum(pmap, d, 0); f = wsum(pmap, d, 1); dl = wsum(pmap, d, 2)
            tot_r.append(int(t)); frd_r.append(int(f)); dlr_r.append(int(dl))
            afpr_r.append(round(t / max(1, f), 1))
            hs_r.append(int(wsum(pmap, d, 3)))
        tier_dlr = {T: [int(wsum(e["tier"][T], d, 0)) for d in axis] for T in ("T1", "T2", "T3")}
        tier_frd = {T: [int(wsum(e["ttot"][T], d, 1)) for d in axis] for T in ("T1", "T2", "T3")}
        tiers = []
        for T in ("T1", "T2", "T3"):
            t = wsum(e["ttot"][T], max_d, 0); f = wsum(e["ttot"][T], max_d, 1); dl = wsum(e["ttot"][T], max_d, 2)
            afpr = round(t / max(1, f), 1)
            tiers.append({"t": T, "afpr": afpr, "dfpr": round(afpr * 0.92, 1),
                          "fdlr": int(dl), "tot": int(t), "frd": int(f)})
        proc = {"id": src, "short": src, "afprReal": afpr_r,
                "totReal": tot_r, "frdReal": frd_r, "dlrReal": dlr_r,
                "tierDlr": tier_dlr, "tierFrd": tier_frd, "tiers": tiers}
        if has_hs:
            proc["dlrRealHs"] = hs_r                                   # else: omitted → dashboard shows "HS data N/A"
        if has_v15:
            proc["tierDlrV15"] = {T: [int(wsum(e["tv15"][T], d, 0)) for d in axis] for T in ("T1", "T2", "T3")}
        procs.append(proc)
    procs.sort(key=lambda p: p["totReal"][-1], reverse=True)
    return {family: {"dates": dates, "procs": procs}}


def build_payload(path, max_rows, max_sheet_rows, validate=False):
    wb = load_workbook(path, read_only=True, data_only=True)
    rule_rows = {}
    sheets = {}
    report = []

    for ws in wb.worksheets:
        headers, data = read_sheet(ws)
        sheets[ws.title] = [
            {k: json_safe(v) for k, v in r.items()}
            for r in data[:max_sheet_rows]
        ]
        view = sheet_to_view(ws.title)
        cols = resolve_columns(headers)
        if view:
            rule_rows[view] = map_rule_rows(data, cols, max_rows)
        report.append((ws.title, view, headers, cols, len(data)))

    if validate:
        print("\n=== VALIDATION REPORT ===")
        print(f"workbook: {os.path.basename(path)}   sheets: {len(wb.worksheets)}\n")
        for title, view, headers, cols, n in report:
            tag = f"-> view '{view}'" if view else "-> (raw sheet only)"
            print(f"• {title!r}  ({n} data rows)  {tag}")
            if view:
                missing = [f for f, h in cols.items() if h is None]
                got = {f: h for f, h in cols.items() if h}
                print(f"      matched: {got}")
                if missing:
                    print(f"      *** MISSING columns: {missing}  (check COLUMN_ALIASES)")
            print()
        covered = [v for v in VIEW_SHEET_RULES if v in rule_rows]
        uncovered = [v for v in VIEW_SHEET_RULES if v not in rule_rows]
        print(f"views populated: {covered or 'NONE'}")
        if uncovered:
            print(f"views with no matching sheet: {uncovered}")
        print("=========================\n")

    return {
        "reportName": os.path.basename(path),
        "generatedAt": dt.datetime.now().strftime("%Y-%m-%d %H:%M"),
        "ruleRows": rule_rows,
        "sheets": sheets,
    }


def inject(template_path, payload, out_path):
    with open(template_path, "r", encoding="utf-8") as f:
        html = f.read()
    if PLACEHOLDER not in html:
        sys.exit(f"ERROR: placeholder not found in template:\n    {PLACEHOLDER}")
    # json.dumps is valid JS-object syntax; embed safely (escape </script>).
    js = "const INJECTED_DATA = " + json.dumps(payload, ensure_ascii=False, default=str) + ";"
    js = js.replace("</", "<\\/")  # avoid premature </script> close
    out = html.replace(PLACEHOLDER, js, 1)
    with open(out_path, "w", encoding="utf-8") as f:
        f.write(out)
    kb = len(out.encode("utf-8")) / 1024
    n_views = len(payload.get("ruleRows", {}))
    n_sheets = len(payload.get("sheets", {}))
    vol = payload.get("volume", {})
    vol_desc = ", ".join("%s:%d" % (g, len(v.get("procs", []))) for g, v in vol.items()) or "—"
    fams = payload.get("families", {})
    fam_desc = ", ".join("%s:%d" % (f, len(v.get("procs", []))) for f, v in fams.items()) or "—"
    print(f"OK  wrote {out_path}  ({kb:.0f} KB)  — rule views {n_views}, sheets {n_sheets}, volume[{vol_desc}], families[{fam_desc}]")


def main():
    ap = argparse.ArgumentParser(description="Assemble CMPT dashboard data (Excel + Snowflake) and inject into the HTML.")
    ap.add_argument("--excel", help="path to the rule-performance .xlsx workbook (optional)")
    ap.add_argument("--snowflake", action="store_true", help="pull CPP batch volume from Snowflake into §04")
    ap.add_argument("--creds", help="path to the cred json (default: %s)" % CRED_PATH)
    ap.add_argument("--template", default="cmpt_report_pro.html", help="HTML template with the placeholder")
    ap.add_argument("--out", help="output HTML path (omit with --validate)")
    ap.add_argument("--validate", action="store_true", help="print what the Excel parse found, write nothing")
    ap.add_argument("--max-rows", type=int, default=5000, help="cap rows per rule view")
    ap.add_argument("--max-sheet-rows", type=int, default=2000, help="cap rows per raw sheet")
    args = ap.parse_args()

    # Stamp the masthead from THIS build run (Python time), so the report shows when it was built
    # — not the viewer's clock when they open the .html. The template reads INJECTED_DATA.meta.
    _build_ts = dt.datetime.now()
    payload = {
        "generatedAt": _build_ts.strftime("%Y-%m-%d %H:%M"),   # kept for back-compat
        "meta": {
            "generated": _build_ts.strftime("%Y-%m-%d %H:%M"),
            # "Data as of" defaults to the build date; point it at the true latest data date if the
            # SQL exposes one (fraud data usually lags ~1 day behind the build).
            "dataAsOf": _build_ts.strftime("%Y-%m-%d"),
        },
    }

    if args.excel:
        payload.update(build_payload(args.excel, args.max_rows, args.max_sheet_rows, validate=args.validate))
    if args.validate:
        return

    if args.snowflake:
        conn = snowflake_connect(args.creds)
        try:
            for group, sql in VOLUME_SQL.items():
                if "<YOUR_" in sql:                       # unfilled placeholder → skip
                    print("  skip %s: SQL not filled in" % group)
                    continue
                rows = fetch_cpp_volume(conn, sql)
                vol = build_volume_payload(rows, group=group)
                if vol:
                    payload.setdefault("volume", {}).update(vol)
                    print("  fetched %s volume: %d sources x %d days" % (group, len(vol[group]["procs"]), len(vol[group]["dates"])))
                else:
                    print("  WARNING: %s volume query returned no rows" % group)
            # Q2 — process daily metrics, one SQL per family (§00–§02, §05)
            for fam, sql in PROCESS_SQL.items():
                if "<YOUR_" in sql or "<PASTE" in sql:        # unfilled placeholder → skip
                    print("  skip %s families: SQL not filled in" % fam)
                    continue
                # roll=1: the SQL already emits per-as-of-day trailing-15d windowed values
                # (distinct accts deduped + summed $), so Python only aligns — no second window.
                fp = build_family_payload(fetch_cpp_volume(conn, sql), fam, roll=1)
                if fp:
                    payload.setdefault("families", {}).update(fp)
                    print("  fetched %s process metrics: %d sources x %d days" % (fam, len(fp[fam]["procs"]), len(fp[fam]["dates"])))
                else:
                    print("  WARNING: %s process-metrics query returned no rows" % fam)
        finally:
            conn.close()

    if not args.excel and not args.snowflake:
        sys.exit("ERROR: provide at least one source — --excel and/or --snowflake")
    if not args.out:
        sys.exit("ERROR: --out is required when not using --validate")
    inject(args.template, payload, args.out)


if __name__ == "__main__":
    main()
